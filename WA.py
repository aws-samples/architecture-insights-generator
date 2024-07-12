import os
import sqlite3
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image, ImageTk
import traceback
import boto3
import datetime
import logging

# 配置日志记录器
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# 创建一个控制台处理器
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)

# 创建一个文件处理器
file_handler = logging.FileHandler('output.log')
file_handler.setLevel(logging.INFO)

# 创建一个格式化器
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

# 为处理器添加格式化器
console_handler.setFormatter(formatter)
file_handler.setFormatter(formatter)

# 为日志记录器添加处理器
logger.addHandler(console_handler)
logger.addHandler(file_handler)


window_width = 600
window_height = window_width - 300
hide_secret = "                                                            "

# 创建Tkinter窗口
root = tk.Tk()
root.title("Architecture Insights Generator")

# 设置窗口大小和位置
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = (screen_width - window_width) // 2
y = (screen_height - window_height) // 2
root.geometry(f"{window_width}x{window_height}+{x}+{y}")

# 加载背景图片
bg_image = Image.open("Chimera.png")
#bg_image = bg_image.resize((window_width, window_height), resample=Image.Resampling.LANCZOS)
bg_photo = ImageTk.PhotoImage(bg_image)

# 创建背景标签
bg_label = tk.Label(root, image=bg_photo)
bg_label.place(x=0, y=0, relwidth=1, relheight=1)


# 创建Label小部件
text_frame = tk.Frame(root, bg="white")
text_frame.place(relx=0.5, rely=0.3, anchor="center", relwidth=0.7, relheight=0.45)
msg_label = tk.Label(text_frame, text="Please import the xlsx file exported from AWS Trusted Advisor into Architecture Insights Generator using the button below.", bg="white", fg="black", wraplength=350, justify="center", height=5)
msg_label.pack(pady=10)

# 创建文件选择按钮
file_button = tk.Button(root, text="Select Excel File", font=("Arial", 14), command=lambda: browse_file(), width=10)
file_button.place(relx=0.5, rely=0.62, anchor="center")

# 创建设置按钮
settings_button = tk.Button(root, text="Settings", font=("Arial", 14), command=lambda: show_settings_window(), width=10)
settings_button.place(relx=0.5, rely=0.74, anchor="center")

# 创建更新WA按钮
update_button = tk.Button(root, text="Update Workload", font=("Arial", 14), command=lambda: update_workload_with_TA(), width=10)
update_button.place(relx=0.5, rely=0.86, anchor="center")
update_button.config(state="disabled")

# 创建设置对象
settings = {
    'region': None,
    'workload': None,
    'lens': None,
    'override_notes' : False
}

# 创建TA检查结果对象
TA_results = {}

# 获取AWS区域列表
session = boto3.Session()
regions = session.get_available_regions('wellarchitected')

if os.path.exists('data.db'):
    # 如果文件存在,则删除它
    os.remove('data.db')
    logger.info('**********Deleted data.db***********')
# 创建SQLite数据库连接
conn = sqlite3.connect('data.db')
c = conn.cursor()


def get_workloads_and_lenses(region):
    wellarchitected = boto3.client('wellarchitected', region_name=region)
    workloads = []
    next_token = ''
    while True:
        if next_token:
            response = wellarchitected.list_workloads(NextToken=next_token)
        else:
            response = wellarchitected.list_workloads()
        workloads.extend([workload['WorkloadName'] + hide_secret + workload['WorkloadId'] for workload in response['WorkloadSummaries']])
        next_token = response.get('NextToken')
        if not next_token:
            break
    lenses = []
    next_token = ''
    while True:
        if next_token:
            response = wellarchitected.list_lenses(NextToken=next_token)
        else:
            response = wellarchitected.list_lenses()
        lenses.extend([lens['LensName'] + hide_secret + lens['LensArn'] for lens in response['LensSummaries']])
        next_token = response.get('NextToken')
        if not next_token:
            break
    
    return workloads, lenses

def show_settings_window():
    global settings_window
    settings_window = tk.Toplevel(root)
    settings_window.title("Settings")
    settings_window.grab_set()  # 确保小窗口获得焦点

    # 获取主窗口的大小和位置
    root_width = root.winfo_width()
    root_height = root.winfo_height()
    root_x = root.winfo_x()
    root_y = root.winfo_y()

    # 计算小窗口的位置
    window_width = 350
    window_height = 200
    window_x = root_x + (root_width - window_width) // 2
    window_y = root_y + (root_height - window_height) // 2

    # 设置小窗口的位置
    settings_window.geometry(f"{window_width}x{window_height}+{window_x}+{window_y}")

    region_var = tk.StringVar()
    workload_var = tk.StringVar()
    lens_var = tk.StringVar()
    override_notes_var = tk.BooleanVar()

    region_label = ttk.Label(settings_window, text="Region:")
    region_dropdown = ttk.Combobox(settings_window, textvariable=region_var, state='readonly')
    region_dropdown['values'] = regions

    workload_label = ttk.Label(settings_window, text="Workload:")
    workload_dropdown = ttk.Combobox(settings_window, textvariable=workload_var, state='disabled')

    lens_label = ttk.Label(settings_window, text="Custom Lens:")
    lens_dropdown = ttk.Combobox(settings_window, textvariable=lens_var, state='disabled')

    override_notes_label = ttk.Label(settings_window, text="Override notes for each question in the workload")
    override_notes_checkbox = ttk.Checkbutton(settings_window, variable=override_notes_var)


    def update_workloads_and_lenses(*args):
        selected_region = region_var.get()
        if selected_region:
            try:
                workloads, lenses = get_workloads_and_lenses(selected_region)
                # workload_dropdown.config(state='readonly', values=['Create a new workload'] + workloads)
                workload_dropdown.config(state='readonly', values=workloads)
                lens_dropdown.config(state='readonly', values=lenses)
            except Exception as e:
                workload_dropdown.config(state='disabled', values=[])
                lens_dropdown.config(state='disabled', values=[])
                logger.exception(f"Error: {e}")
                traceback.print_exc()
        else:
            workload_dropdown.config(state='disabled', values=[])
            lens_dropdown.config(state='disabled', values=[])

    def save_settings():
        global settings
        settings = {
            'region': region_var.get(),
            'workload': workload_var.get(),
            'lens': lens_var.get(),
            'override_notes': override_notes_var.get()
        }
        logger.info(settings)
        # 在这里可以将设置保存到文件或数据库中
        settings_window.destroy()

    save_button = ttk.Button(settings_window, text="Save", command=save_settings)

    # 设置 region 下拉框的默认值
    if 'region' in settings and settings['region'] != None:
        region_var.set(settings['region'])
    # 设置 workload 和 lens 下拉框的默认值
    if 'workload' in settings and settings['workload'] != None:
        workload_var.set(settings['workload'])
        workload_dropdown.config(state='readonly')
    if 'lens' in settings and settings['lens'] != None:
        lens_var.set(settings['lens'])
        lens_dropdown.config(state='readonly')
    if 'override_notes' in settings:
        override_notes_var.set(settings['override_notes'])

    region_var.trace('w', update_workloads_and_lenses)

    region_label.grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
    region_dropdown.grid(row=0, column=1, padx=5, pady=5)
    workload_label.grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
    workload_dropdown.grid(row=1, column=1, padx=5, pady=5)
    lens_label.grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
    lens_dropdown.grid(row=2, column=1, padx=5, pady=5)

    # 以下被注释的代码功能是提供一个checkbox，提供一个选项，允许追加notes而不是覆盖notes，但因为wa 的notes有长度限制，当check项目问题较多时追加notes会引起报错，所以谨慎起见，注释掉，不提供追加的功能。
    # override_notes_label.grid(row=3, column=1, padx=5, pady=5, sticky=tk.W)
    # override_notes_checkbox.grid(row=3, column=0, padx=5, pady=5, sticky=tk.W)

    #save_button.grid(row=3, columnspan=2, padx=5, pady=8)
    save_button.place(relx=0.5, rely=0.82, anchor="center")


# 创建文件选择框
def browse_file():
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select Excel File", filetypes=(("Excel Files", "*.xlsx"),))
    if filename:
        import_excel(filename)

# 导入Excel文件到SQLite
def import_excel(filename):
    file_name = os.path.basename(filename).split('.')[0]
    table_name = "TA_all"
    
    # 删除同名表
    
    # 创建主表
    create_table_query = "CREATE TABLE IF NOT EXISTS " + table_name + " (check_index TEXT PRIMARY KEY, check_item TEXT, account_id TEXT, description TEXT, status TEXT)"
    c.execute(create_table_query)
    
    # 读取Excel文件
    xlsx = pd.ExcelFile(filename)
    
    # 循环每个sheet页
    for sheet_name in xlsx.sheet_names:
        df = xlsx.parse(sheet_name, header=None)
        
        # 提取标题和账号ID
        check_item = df.iloc[0, 0]
        account_id = df.iloc[1, 0].split(': ')[1]
        description = df.iloc[2, 0].split(': ')[1]
        status = df.iloc[3, 0]
        
        # 插入数据
        insert_query = "INSERT OR REPLACE INTO " + table_name + " (check_index, check_item, account_id, description, status) VALUES (?, ?, ?, ?, ?)"
        c.execute(insert_query, (sheet_name, check_item, account_id, description, status))

        df_check = xlsx.parse(sheet_name)
        if df_check.shape[0] > 9:
            try:
                df_detail = xlsx.parse(sheet_name, header=9)
                df_detail.to_sql(sheet_name, conn, if_exists='replace', index=False)
            except Exception as e:
                # 打印异常信息
                #logger.info("An exception occurred:")
                logger.exception(f"Error: {e}")
                # 打印异常的堆栈跟踪
                #logger.info("Traceback:")
                #traceback.print_exc()

    # 创建Lens表
    table_name = 'lens'
    c.execute(f'DROP TABLE IF EXISTS {table_name}')

    with open('output.csv', 'r') as csvfile:
        reader = csv.reader(csvfile)
        columns = next(reader)  # 获取列名
        columns_str = ', '.join([f'"{col}"' for col in columns])
        c.execute(f'CREATE TABLE {table_name} ({columns_str} TEXT)')

        # 插入数据
        insert_query = f'INSERT INTO {table_name} VALUES ({",".join(["?"] * len(columns))})'
        c.executemany(insert_query, reader)

    # 提交更改
    conn.commit()

    output_excel()

    messagebox.showinfo("Success", f"'{filename}' analysis successful")
    update_msg_label("The analysis has been saved in the TA-check.xlsx file in the current directory. If you need to update it to the Notes in the AWS WA Tool, please first set the necessary parameters through the 'Settings'.")

    update_button.config(state="normal")



def output_excel():
    # 连接SQLite数据库
    conn = sqlite3.connect('data.db')
    c = conn.cursor()

    # 执行SQL查询
    query = """
    SELECT a.check_index, b.[Question ID], b.[Choice ID], b.[Pillar Name], b.[Question Title], b.[Choice Title], b.[Trusted Advisor Checks], a.description
    FROM TA_all a, lens b
    WHERE b.[Trusted Advisor Checks] LIKE a.check_item||'%'
      AND a.status IN ('Status: warning', 'Status: error')
    ORDER BY b.[Pillar Name], b.[Question Title];
    """
    c.execute(query)
    results = c.fetchall()
    global TA_results
    TA_results = results.copy()

    # 将结果转换为DataFrame
    columns = [desc[0] for desc in c.description][3:]  # 排除check_index, Question ID, Choice ID列
    df = pd.DataFrame([row[3:] for row in results], columns=columns)

    # 创建新的Excel文件
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "TA-check"

    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    for col in range(df.shape[1]):
        cell = worksheet.cell(row=1, column=col + 1)
        cell.value = df.columns[col]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    worksheet.column_dimensions["A"].width = 20
    worksheet.column_dimensions["B"].width = 60
    worksheet.column_dimensions["C"].width = 60
    worksheet.column_dimensions["D"].width = 60
    worksheet.column_dimensions["E"].width = 60

    # 写入数据
    for row in range(df.shape[0]):
        for col in range(df.shape[1]):
            cell = worksheet.cell(row=row + 2, column=col + 1)
            cell.value = df.iloc[row, col]

    # 循环结果集,创建新的工作表
    for check_index in [row[0] for row in results]:
        try:
            query = f"SELECT * FROM [{check_index}];"
            logger.info("********Detail table SQL************")
            logger.info(query)
            c.execute(query)
        except Exception as e:
            # 打印异常信息
            #logger.info("An exception occurred:")
            logger.exception(f"Error: {e}")
            # 打印异常的堆栈跟踪
            #logger.info("Traceback:")
            #traceback.print_exc()
            continue

        detail_results = c.fetchall()
        detail_columns = [desc[0] for desc in c.description]
        detail_df = pd.DataFrame(detail_results, columns=detail_columns)

        worksheet = workbook.create_sheet(check_index[:29])

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        for col, column_name in enumerate(detail_columns, start=1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = column_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        detail_df = pd.DataFrame(detail_results, columns=detail_columns)


        # 自动调整列宽并写入数据
        for col in range(detail_df.shape[1]):
            column_letter = get_column_letter(col + 1)
            try:
                detail_df_numeric = detail_df.iloc[:, col].apply(pd.to_numeric, errors='coerce')
                detail_df_numeric = detail_df_numeric.dropna()
                max_value = detail_df_numeric.max()
                worksheet.column_dimensions[column_letter].width = max(len(str(max_value)) * 1.2, 20)
            except ValueError:
                worksheet.column_dimensions[column_letter].width = 20


        for row in range(detail_df.shape[0]):
            for col in range(detail_df.shape[1]):
                cell = worksheet.cell(row=row + 2, column=col + 1)
                cell.value = detail_df.iloc[row, col]


    # 保存Excel文件
    workbook.save("TA-check.xlsx")

    # 关闭数据库连接
    conn.close()

def update_workload_with_TA():
    workload_id = ''
    lens_alias = ''
    region_id = ''
    update_count = 0
    overide_notes = True
    try:
        region_id = settings['region']
        workload_id = settings['workload'].split(hide_secret)[-1]
        lens_alias = settings['lens'].split(hide_secret)[-1]
        overide_notes = settings['override_notes']
    except Exception as e:
        logger.exception(f"Error: {e}")
        traceback.print_exc()
        messagebox.showerror("ERROR", "Your settings is not configured correctly, please check.")
        return


    wellarchitected_client = boto3.client('wellarchitected')
    """
        SELECT a.check_index, b.[Question ID], b.[Choice ID], b.[Pillar Name], b.[Question Title], b.[Choice Title], b.[Trusted Advisor Checks], a.description
        FROM TA_all a, lens b
        WHERE b.[Trusted Advisor Checks] LIKE a.check_item||'%'
          AND a.status IN ('Status: warning', 'Status: error')
        ORDER BY b.[Pillar Name], b.[Question Title];
    """
    # 遍历问题数组并更新答案
    TA_results_size = len(TA_results)
    for question in TA_results:
        try:
            logger.info('Updating:'+workload_id+','+lens_alias+','+question[1])
            notes = ''

            response = wellarchitected_client.get_answer(
                WorkloadId=workload_id,
                LensAlias=lens_alias,
                QuestionId=question[1]
                )
            if 'Notes' in response['Answer']:
                notes = response['Answer']['Notes']

            notes = '\n--Updated by Architecture Insights Generator at ' + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + '--\n' + question[5] + '\n' + question[6] + '\n-- --\n' + notes

            response = wellarchitected_client.update_answer(
                WorkloadId=workload_id,
                LensAlias=lens_alias,
                QuestionId=question[1],
                Notes=notes[:2000]
            )
            update_count += 1
            logger.info(f"Successfully updated answer for question: {question[1]}")
            logger.info(f"Updating...{round(100*update_count/TA_results_size)}%")

        except Exception as e:
            logger.info(f"Error at:{question[1]}")
            logger.exception(f"Error: {e}")
            traceback.print_exc()

    messagebox.showinfo("Success", f"Updated {update_count} answer(s). ")
    update_msg_label(f"There were {TA_results_size} questions in total, and {update_count} were successfully updated.")


# 定义更新Label文本的函数
def update_msg_label(msg):
    global msg_label
    msg_label.config(text=msg)



root.mainloop()

# 关闭数据库连接
conn.close()
